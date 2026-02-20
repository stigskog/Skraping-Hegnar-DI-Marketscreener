import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)

# === Stox.no Color Palette (from output.xlsx) ===
COLOR_TITLE_BG = 'FFF8E7'        # Cream background for title
COLOR_TITLE_TEXT = '1B4332'       # Dark green text
COLOR_SECTION_BG = '2D6A4F'      # Medium green for country section headers
COLOR_SECTION_TEXT = 'FFFFFF'     # White text
COLOR_HEADER_BG = '1B4332'       # Dark green for column headers
COLOR_HEADER_TEXT = 'FFFFFF'      # White text
COLOR_ROW_EVEN = 'FFF8E7'        # Cream for even rows
COLOR_ROW_ODD = 'F5F0E1'         # Slightly darker cream for odd rows
COLOR_DATA_BOLD = '1B4332'       # Dark green for bold data text
COLOR_DATA_NORMAL = '333333'     # Dark gray for normal text
COLOR_BEARISH = '9C0006'         # Red for bearish
COLOR_NEUTRAL = '888888'         # Gray for 0% change

# Fills
FILL_TITLE = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type='solid')
FILL_SECTION = PatternFill(start_color=COLOR_SECTION_BG, end_color=COLOR_SECTION_BG, fill_type='solid')
FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid')
FILL_ROW_EVEN = PatternFill(start_color=COLOR_ROW_EVEN, end_color=COLOR_ROW_EVEN, fill_type='solid')
FILL_ROW_ODD = PatternFill(start_color=COLOR_ROW_ODD, end_color=COLOR_ROW_ODD, fill_type='solid')

# Fonts
FONT_TITLE = Font(name='Arial', size=14, bold=True, color=COLOR_TITLE_TEXT)
FONT_SECTION = Font(name='Arial', size=11, bold=True, color=COLOR_SECTION_TEXT)
FONT_HEADER = Font(name='Arial', size=10, bold=True, color=COLOR_HEADER_TEXT)
FONT_DATA_BOLD = Font(name='Arial', size=9, bold=True, color=COLOR_DATA_BOLD)
FONT_DATA_NORMAL = Font(name='Arial', size=9, bold=False, color=COLOR_DATA_NORMAL)
FONT_BEARISH_BOLD = Font(name='Arial', size=9, bold=True, color=COLOR_BEARISH)
FONT_BEARISH_NORMAL = Font(name='Arial', size=9, bold=False, color=COLOR_BEARISH)
FONT_NEUTRAL = Font(name='Arial', size=9, bold=False, color=COLOR_NEUTRAL)

# Alignment
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

# Country config
COUNTRY_SECTIONS = [
    ('NO', '\U0001f1f3\U0001f1f4 Norske aksjer (Oslo B\u00f8rs)'),
    ('SE', '\U0001f1f8\U0001f1ea Svenske aksjer (Stockholmsb\u00f6rsen)'),
    ('DK', '\U0001f1e9\U0001f1f0 Danske aksjer (K\u00f8benhavn)'),
    ('FI', '\U0001f1eb\U0001f1ee Finske aksjer (Helsinki)'),
    ('US', '\U0001f1fa\U0001f1f8 Amerikanske aksjer (USA)'),
]


def get_pct_font(pct_str, is_bearish=False):
    """Return appropriate font for percentage value."""
    if not pct_str:
        return FONT_DATA_BOLD
    clean = pct_str.replace('%', '').replace('+', '').replace(',', '.').strip()
    try:
        val = float(clean)
        if val < 0:
            return Font(name='Arial', size=9, bold=True, color=COLOR_BEARISH)
        elif val == 0:
            return FONT_NEUTRAL
        else:
            return FONT_DATA_BOLD
    except:
        return FONT_DATA_BOLD


def generate_excel(signals, output_path=None):
    """Generate the Excel file matching the exact Stox.no format.

    signals: list of dicts with keys:
        company_name, ticker, direction, comment, time, country, pct_change (optional)
    """
    now = datetime.now()
    days_no = ['Mandag', 'Tirsdag', 'Onsdag', 'Torsdag', 'Fredag', 'L\u00f8rdag', 'S\u00f8ndag']
    months_no = ['', 'januar', 'februar', 'mars', 'april', 'mai', 'juni',
                 'juli', 'august', 'september', 'oktober', 'november', 'desember']
    day_name = days_no[now.weekday()]
    date_str = f"{now.day}. {months_no[now.month]}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Signaler {now.day}. {months_no[now.month][:3]}"

    # Column widths (matching output.xlsx)
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 52
    ws.column_dimensions['E'].width = 3  # spacer
    ws.column_dimensions['F'].width = 56
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 10

    row = 1

    # === Title Row ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    cell_a = ws.cell(row=row, column=1, value=f'Nyheter aksjer \u2013 {day_name} {date_str} {now.year}')
    cell_a.font = FONT_TITLE
    cell_a.fill = FILL_TITLE
    cell_a.alignment = ALIGN_LEFT
    cell_f = ws.cell(row=row, column=6, value='Sortert etter aksjenavn')
    cell_f.font = FONT_TITLE
    cell_f.fill = FILL_TITLE
    cell_f.alignment = ALIGN_LEFT
    # Fill remaining title cells
    for col in range(1, 10):
        c = ws.cell(row=row, column=col)
        c.fill = FILL_TITLE
    ws.row_dimensions[row].height = 42
    row += 1

    # === Spacer row ===
    ws.row_dimensions[row].height = 4
    row += 1

    # Group signals by country
    country_signals = {}
    for sig in signals:
        country = sig.get('country', 'NO')
        if country not in country_signals:
            country_signals[country] = []
        country_signals[country].append(sig)

    # Generate sections for each country
    for country_code, section_title in COUNTRY_SECTIONS:
        sigs = country_signals.get(country_code, [])
        if not sigs:
            continue

        # Sort: left side by time (descending), right side by company name (ascending)
        by_time = sorted(sigs, key=lambda s: s.get('time', ''), reverse=True)
        by_name = sorted(sigs, key=lambda s: (s.get('company_name', '') + ' ' + s.get('ticker', '')).lower())

        # === Country Section Header ===
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
        cell_sec_a = ws.cell(row=row, column=1, value=f'  {section_title}')
        cell_sec_a.font = FONT_SECTION
        cell_sec_a.fill = FILL_SECTION
        cell_sec_a.alignment = ALIGN_LEFT
        cell_sec_f = ws.cell(row=row, column=6, value=f'  {section_title}')
        cell_sec_f.font = FONT_SECTION
        cell_sec_f.fill = FILL_SECTION
        cell_sec_f.alignment = ALIGN_LEFT
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = FILL_SECTION
        ws.row_dimensions[row].height = 26
        row += 1

        # === Column Headers ===
        headers_left = ['Sortert etter kl.slett', 'Retning', '%', 'Kommentar']
        headers_right = ['Sortert etter aksjenavn / ticker', 'Retning', '%', 'Kl.']
        for i, h in enumerate(headers_left):
            c = ws.cell(row=row, column=i + 1, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_LEFT
        for i, h in enumerate(headers_right):
            c = ws.cell(row=row, column=i + 6, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_LEFT
        ws.cell(row=row, column=5).fill = PatternFill()  # spacer stays empty
        ws.row_dimensions[row].height = 26
        row += 1

        # === Data Rows ===
        max_rows = max(len(by_time), len(by_name))
        for r_idx in range(max_rows):
            row_fill = FILL_ROW_EVEN if r_idx % 2 == 0 else FILL_ROW_ODD

            # Left side (sorted by time)
            if r_idx < len(by_time):
                sig = by_time[r_idx]
                is_bearish = sig.get('direction', '').lower() == 'bearish'
                time_val = sig.get('time', '')
                ticker = sig.get('ticker', '')
                company = sig.get('company_name', '')
                pct = sig.get('pct_change', '')
                comment = sig.get('comment', '')

                # Column A: time + company (ticker)
                a_val = f"{time_val} \u2013 {company} ({ticker})" if time_val else f"{company} ({ticker})"
                c_a = ws.cell(row=row, column=1, value=a_val)
                c_a.font = FONT_DATA_BOLD
                c_a.fill = row_fill
                c_a.alignment = ALIGN_LEFT

                # Column B: direction
                dir_text = '\u25cf Bullish' if not is_bearish else '\u25cf Bearish'
                c_b = ws.cell(row=row, column=2, value=dir_text)
                c_b.font = FONT_BEARISH_BOLD if is_bearish else FONT_DATA_BOLD
                c_b.fill = row_fill
                c_b.alignment = ALIGN_LEFT

                # Column C: percentage
                c_c = ws.cell(row=row, column=3, value=pct)
                c_c.font = get_pct_font(pct, is_bearish)
                c_c.fill = row_fill
                c_c.alignment = ALIGN_LEFT

                # Column D: comment
                c_d = ws.cell(row=row, column=4, value=comment)
                c_d.font = FONT_DATA_NORMAL
                c_d.fill = row_fill
                c_d.alignment = ALIGN_LEFT
            else:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = row_fill

            # Column E: spacer (always empty)

            # Right side (sorted by company name)
            if r_idx < len(by_name):
                sig = by_name[r_idx]
                is_bearish = sig.get('direction', '').lower() == 'bearish'
                time_val = sig.get('time', '')
                ticker = sig.get('ticker', '')
                company = sig.get('company_name', '')
                pct = sig.get('pct_change', '')
                comment = sig.get('comment', '')

                # Column F: company (ticker): comment
                f_val = f"{company} ({ticker}): {comment}"
                c_f = ws.cell(row=row, column=6, value=f_val)
                c_f.font = FONT_DATA_BOLD
                c_f.fill = row_fill
                c_f.alignment = ALIGN_LEFT

                # Column G: direction
                dir_text = '\u25cf Bullish' if not is_bearish else '\u25cf Bearish'
                c_g = ws.cell(row=row, column=7, value=dir_text)
                c_g.font = FONT_BEARISH_BOLD if is_bearish else FONT_DATA_BOLD
                c_g.fill = row_fill
                c_g.alignment = ALIGN_LEFT

                # Column H: percentage
                c_h = ws.cell(row=row, column=8, value=pct)
                c_h.font = get_pct_font(pct, is_bearish)
                c_h.fill = row_fill
                c_h.alignment = ALIGN_LEFT

                # Column I: time
                c_i = ws.cell(row=row, column=9, value=time_val)
                c_i.font = FONT_DATA_NORMAL
                c_i.fill = row_fill
                c_i.alignment = ALIGN_LEFT
            else:
                for col in range(6, 10):
                    ws.cell(row=row, column=col).fill = row_fill

            ws.row_dimensions[row].height = 26
            row += 1

        # Spacer row between sections
        ws.row_dimensions[row].height = 4
        row += 1

    # Remove trailing empty spacer
    if row > 1:
        row -= 1

    # Save
    if output_path is None:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'outputs')
        os.makedirs(output_dir, exist_ok=True)
        filename = f"signaler_{now.strftime('%Y-%m-%d_%H%M')}.xlsx"
        output_path = os.path.join(output_dir, filename)

    wb.save(output_path)
    logger.info(f"Excel saved to {output_path}")
    return output_path
